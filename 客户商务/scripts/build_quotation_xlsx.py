#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成 Excel 报价至 客户商务/碳足迹系统_模块化报价单.xlsx。

**单工作表**：按应用模块（M01–M10）与功能点列示；**统一 1500 元/人天**，不按角色拆分单价。
SpreadJS 为固定价单列。人天均为整数。
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "碳足迹系统_模块化报价单.xlsx"

# --- 统一人天单价（元/人天）与总价闭合 ---
LABOR_RATE = 1500
# 人工 466 人天 × 1500 = 699000 元；SpreadJS 固定费 101000 元；合计 800000 元
SPREADJS_FIXED_YUAN = 101_000
TOTAL_QUOTE_YUAN = 800_000
MVP_LABOR_DAYS = 333  # 第一阶段（MVP）人天（不含 SpreadJS）
PHASE2_OPT_LABOR_DAYS = 133  # 二阶段 + 可选 人天

assert (
    (MVP_LABOR_DAYS + PHASE2_OPT_LABOR_DAYS) * LABOR_RATE + SPREADJS_FIXED_YUAN
    == TOTAL_QUOTE_YUAN
)


def _split_int(total_days: int, n: int) -> list:
    """将 total_days 拆成 n 个非负整数，尽可能均分（大数在前）。"""
    if n <= 0:
        return []
    base = total_days // n
    rem = total_days % n
    return [base + (1 if i < rem else 0) for i in range(n)]


# 报价编号 → 应用模块（与产品建设方案总册 M01–M10 一致）
QID_MXX = {
    "Q-PLAT-001": "M09",
    "Q-PLAT-002": "M09",
    "Q-PLAT-003": "M09",
    "Q-PORT-001": "M01",
    "Q-PORT-002": "M01",
    "Q-OPS-001": "M02",
    "Q-OPS-002": "M02",
    "Q-OPS-003": "M02",
    "Q-OPS-004": "M02",
    "Q-OPS-005": "M02",
    "Q-OPS-006": "M02",
    "Q-OPS-007": "M02",
    "Q-OPS-008": "M02",
    "Q-OPS-009": "M07",
    "Q-OPS-010": "M08",
    "Q-OPS-011": "M08",
    "Q-OPS-012": "M03",
    "Q-OPS-013": "M02",
    "Q-OPS-014": "M02",
    "Q-SUP-001": "M02",
    "Q-SUP-002": "M02",
    "Q-SUP-003": "M03",
    "Q-SUP-004": "M04",
    "Q-SUP-005": "M07",
    "Q-SUP-006": "M08",
    "Q-SUP-007": "M02",
    "Q-SUP-008": "M02",
    "Q-CER-001": "M06",
    "Q-CER-002": "M06",
    "Q-CER-003": "M06",
    "Q-CER-004": "M07",
    "Q-CER-005": "M07",
    "Q-CER-006": "M06",
    "Q-CER-007": "M09",
    "Q-X-001": "M03",
    "Q-X-002": "M03",
    "Q-X-003": "M04",
    "Q-X-004": "M09",
    "Q-INT-001": "M10",
    "Q-INT-002": "M05",
    "Q-INT-003": "M10",
    "Q-INT-004": "M10",
    "Q-INT-005": "M10",
    "Q-STD-001": "M10",
    "Q-PM-001": "M09",
    "Q-PM-002": "M09",
}

MXX_TITLE = {
    "M01": "M01 门户与品牌入口",
    "M02": "M02 订单与任务编排",
    "M03": "M03 模板与结构化采集",
    "M04": "M04 协同与澄清",
    "M05": "M05 LCA 与计算衔接",
    "M06": "M06 核查与认证协作",
    "M07": "M07 报告与证书全周期",
    "M08": "M08 生态主数据",
    "M09": "M09 平台治理与安全基座",
    "M10": "M10 连接与体系互操作",
}


def _build_day_alloc() -> dict[str, int]:
    mvp = [r for r in DETAIL_ROWS if r[1] == "MVP" and r[0] != "Q-X-002"]
    p2 = [r for r in DETAIL_ROWS if r[1] in ("二阶段", "可选")]
    assert len(mvp) == 38 and len(p2) == 7
    assert MVP_LABOR_DAYS + PHASE2_OPT_LABOR_DAYS == 466
    days: dict[str, int] = {}
    for q, d in zip([r[0] for r in mvp], _split_int(MVP_LABOR_DAYS, len(mvp))):
        days[q] = d
    for q, d in zip([r[0] for r in p2], _split_int(PHASE2_OPT_LABOR_DAYS, len(p2))):
        days[q] = d
    days["Q-X-002"] = 1
    return days


def _labor_yuan(day_map: dict[str, int]) -> int:
    s = 0
    for r in DETAIL_ROWS:
        qid = r[0]
        if qid == "Q-X-002":
            continue
        s += LABOR_RATE * day_map[qid]
    return s


# 明细行：(报价编号, 阶段, 一级模块, 二级模块, 功能点, 交付物, 单位, 数量, 单价, 备注, 验收准则)
# 数量/单价留空则金额为公式自动；固定价可填数量与单价
DETAIL_ROWS = [
    # 平台与基础
    (
        "Q-PLAT-001",
        "MVP",
        "平台与基础",
        "RBAC与系统管理",
        "角色权限矩阵与系统管理",
        "Web管理端、权限API、联调",
        "人天",
        None,
        None,
        "范围见 docs/01_业务与流程/02_全局数据字典与枚举.md §3",
        "角色可按矩阵配置关键菜单与操作；权限变更生效可验证",
    ),
    (
        "Q-PLAT-002",
        "MVP",
        "平台与基础",
        "RBAC与系统管理",
        "关键操作日志",
        "日志采集、查询、导出接口与页面",
        "人天",
        None,
        None,
        "",
        "填报/下发/归档等关键动作可追溯且可检索",
    ),
    (
        "Q-PLAT-003",
        "MVP",
        "平台与基础",
        "基础服务",
        "文件存储与上传策略",
        "对象存储对接或自建、凭证与大小限制",
        "人天",
        None,
        None,
        "与模板填报、证照上传共用",
        "凭证可按任务维度上传、下载权限受控",
    ),
    # 门户
    (
        "Q-PORT-001",
        "MVP",
        "门户与统一入口",
        "系统门户",
        "index.html 平台门户与多服务入口",
        "静态页、导航、部署配置",
        "人天",
        None,
        None,
        "见 README 门户说明",
        "可访问且产品碳足迹入口跳转正确",
    ),
    (
        "Q-PORT-002",
        "MVP",
        "门户与统一入口",
        "产品碳足迹门户",
        "pcf.html 二级门户与三端入口",
        "静态页、工作台跳转",
        "人天",
        None,
        None,
        "",
        "运营/供应商/核查入口可分别进入对应工作台",
    ),
    # 运营端 — 第一阶段设计方案 §3.1
    (
        "Q-OPS-001",
        "MVP",
        "运营管理端",
        "运营驾驶舱",
        "简版 KPI、数据流、快捷入口",
        "operator/dashboard.html 及接口",
        "人天",
        None,
        None,
        "整版驾驶舱见「二阶段」独立行",
        "展示核心指标与跳转任务/报告入口",
    ),
    (
        "Q-OPS-002",
        "MVP",
        "运营管理端",
        "订单管理",
        "订单四态与「确认并分配任务」",
        "order.html、订单API、生成任务",
        "人天",
        None,
        None,
        "与鑫采商城对接单列集成行",
        "已支付→确认后生成任务并可从任务列表追溯",
    ),
    (
        "Q-OPS-003",
        "MVP",
        "运营管理端",
        "任务管理",
        "任务列表（五段）与阶段跳转",
        "self_operated_task_list 等、列表API",
        "人天",
        None,
        None,
        "阶段4仅跳转报告管理",
        "列表阶段与任务主状态一致",
    ),
    (
        "Q-OPS-004",
        "MVP",
        "运营管理端",
        "任务管理",
        "任务详情-配置阶段",
        "task_detail_config、模板选择、Snapshot",
        "人天",
        None,
        None,
        "见 05_模板引擎",
        "可选模板并生成快照供填报使用",
    ),
    (
        "Q-OPS-005",
        "MVP",
        "运营管理端",
        "任务管理",
        "任务详情-采集审核",
        "task_detail_collect、审核通过/驳回",
        "人天",
        None,
        None,
        "与供应商填报联动",
        "可审核供应商提交并产生驳回原因",
    ),
    (
        "Q-OPS-006",
        "MVP",
        "运营管理端",
        "任务管理",
        "任务详情-LCA计算",
        "task_detail_lca、状态与手工/接口预留",
        "人天",
        None,
        None,
        "计算引擎对接见集成行",
        "计算阶段状态可推进并记录",
    ),
    (
        "Q-OPS-007",
        "MVP",
        "运营管理端",
        "任务管理",
        "任务详情-核查协作",
        "task_detail_verify、与核查端状态衔接",
        "人天",
        None,
        None,
        "",
        "核查结果可反映到任务与报告流程",
    ),
    (
        "Q-OPS-008",
        "MVP",
        "运营管理端",
        "任务管理",
        "任务工作台",
        "task_workspace、跨子任务协作",
        "人天",
        None,
        None,
        "视实施范围可合并计价",
        "运营可在工作台完成分配的操作",
    ),
    (
        "Q-OPS-009",
        "MVP",
        "运营管理端",
        "报告管理",
        "用印版、下发、申诉处理、归档",
        "report_mgt、report_detail、规则见 08_PR",
        "人天",
        None,
        None,
        "申诉中不可归档等规则需实现",
        "可走通：上传→下发→供应商确认/申诉→归档闭环",
    ),
    (
        "Q-OPS-010",
        "MVP",
        "运营管理端",
        "供应商管理",
        "供应商CRM与详情、证照视图",
        "supplier_mgt、supplier_mgt_detail",
        "人天",
        None,
        None,
        "与供应商 identity 数据对应",
        "机构与证照信息与供应商端一致",
    ),
    (
        "Q-OPS-011",
        "MVP",
        "运营管理端",
        "核查机构管理",
        "核查机构与策略维护",
        "certifier_mgt、certifier_detail",
        "人天",
        None,
        None,
        "",
        "机构可派发核查任务",
    ),
    (
        "Q-OPS-012",
        "MVP",
        "运营管理端",
        "模板中心",
        "采集模板维护与任务绑定",
        "templates_mgt、template_detail",
        "人天",
        None,
        None,
        "见 05_模板引擎解析逻辑",
        "模板版本可管理并用于生成 Snapshot",
    ),
    (
        "Q-OPS-013",
        "二阶段",
        "运营管理端",
        "结算中心",
        "结算单与确认（完整实现）",
        "settlement 业务与接口",
        "人天",
        None,
        None,
        "第一阶段仅预留入口",
        "按二期范围验收",
    ),
    (
        "Q-OPS-014",
        "二阶段",
        "运营管理端",
        "运营驾驶舱",
        "驾驶舱完整版（整体）",
        "dashboard 增强与数据接入",
        "人天",
        None,
        None,
        "第一阶段设计方案：二迭代做整体",
        "按二期范围验收",
    ),
    # 供应商 — docs/06
    (
        "Q-SUP-001",
        "MVP",
        "供应商工作台",
        "工作台",
        "概览与快捷入口",
        "supplier/dashboard.html",
        "人天",
        None,
        None,
        "见 06 功能清单序号1",
        "待办/澄清/报告数展示正确",
    ),
    (
        "Q-SUP-002",
        "MVP",
        "供应商工作台",
        "待办任务",
        "列表与跳转填报/澄清",
        "task_list.html",
        "人天",
        None,
        None,
        "见 06 序号2",
        "三类待办类型可分流",
    ),
    (
        "Q-SUP-003",
        "MVP",
        "供应商工作台",
        "任务填报",
        "填报与凭证上传",
        "task_fill.html、Spread/表单",
        "人天",
        None,
        None,
        "见 06 序号3",
        "可提交且运营端可看到",
    ),
    (
        "Q-SUP-004",
        "MVP",
        "供应商工作台",
        "待澄清处理",
        "回复与佐证上传",
        "task_clarify.html",
        "人天",
        None,
        None,
        "见 06 序号4",
        "回复与运营端协作衔接",
    ),
    (
        "Q-SUP-005",
        "MVP",
        "供应商工作台",
        "我的报告",
        "预览、确认接收、申诉",
        "reports.html",
        "人天",
        None,
        None,
        "见 06 序号5",
        "仅下发后可见；确认后不可申诉",
    ),
    (
        "Q-SUP-006",
        "MVP",
        "供应商工作台",
        "企业信息与证照",
        "维护/提交空间同步",
        "identity.html",
        "人天",
        None,
        None,
        "见 06 序号6",
        "营业执照必选；同步路径可演示",
    ),
    (
        "Q-SUP-007",
        "MVP",
        "供应商工作台",
        "知识库与帮助",
        "帮助与知识库入口",
        "help 等",
        "人天",
        None,
        None,
        "可与占位内容起价",
        "帮助入口可用",
    ),
    (
        "Q-SUP-008",
        "可选",
        "供应商工作台",
        "绿色生态市场",
        "市场占位扩展为实功能",
        "market.html",
        "人天",
        None,
        None,
        "06 序号7 为 P2",
        "按可选包验收",
    ),
    # 核查 — docs/09
    (
        "Q-CER-001",
        "MVP",
        "认证核查机构工作台",
        "核查仪表盘",
        "KPI与任务队列等",
        "certifier/dashboard.html",
        "人天",
        None,
        None,
        "见 09 序号1",
        "仪表盘可加载并跳转列表",
    ),
    (
        "Q-CER-002",
        "MVP",
        "认证核查机构工作台",
        "核查任务列表",
        "列表与跳转详情",
        "certifier/task_list.html",
        "人天",
        None,
        None,
        "见 09 序号2",
        "taskId 与运营端一致",
    ),
    (
        "Q-CER-003",
        "MVP",
        "认证核查机构工作台",
        "核查任务详情",
        "证照与报告只读、通过/驳回/澄清",
        "certifier/task_detail*.html",
        "人天",
        None,
        None,
        "见 09 序号3；穿透可为简化版",
        "核查动作可驱动任务状态",
    ),
    (
        "Q-CER-004",
        "MVP",
        "认证核查机构工作台",
        "证书颁发管理",
        "证书记录列表与查看",
        "certifier/certificates.html",
        "人天",
        None,
        None,
        "见 09 序号4",
        "与报告/任务可追溯",
    ),
    (
        "Q-CER-005",
        "二阶段",
        "认证核查机构工作台",
        "结算中心",
        "结算单列表与确认",
        "certifier/settlement.html 完整实现",
        "人天",
        None,
        None,
        "09 序号5 为 P1",
        "按二期范围验收",
    ),
    (
        "Q-CER-006",
        "二阶段",
        "认证核查机构工作台",
        "机构管理",
        "人员与成员管理",
        "certifier/admin.html",
        "人天",
        None,
        None,
        "09 序号6 为 P1",
        "按二期范围验收",
    ),
    (
        "Q-CER-007",
        "MVP",
        "认证核查机构工作台",
        "身份与登录",
        "正式环境登录与权限接入",
        "登录页对接 IAM",
        "人天",
        None,
        None,
        "原型 login 为演示；实施期对接",
        "核查员仅见本机构数据",
    ),
    # 横向能力
    (
        "Q-X-001",
        "MVP",
        "横向能力",
        "模板与表格引擎",
        "模板解析、Snapshot、填报渲染",
        "后端解析、Spread 集成、spreadUtils",
        "人天",
        None,
        None,
        "见 docs/02/05_模板引擎解析逻辑.md；授权费另计",
        "与运营模板中心、供应商填报一致",
    ),
    (
        "Q-X-002",
        "MVP",
        "横向能力",
        "第三方费用",
        "GrapeCity SpreadJS 商业授权",
        "License 采购（按年/部署域）",
        "固定价",
        1,
        None,
        "单价填厂商报价；或删行改由客户自采",
        "取得合法授权后生产可用",
    ),
    (
        "Q-X-003",
        "MVP",
        "横向能力",
        "澄清与消息",
        "驳回/澄清/异议与消息可见性",
        "UI+消息API+联调，见 docs/04",
        "人天",
        None,
        None,
        "含钉钉式澄清区若纳入范围",
        "三端消息与状态一致",
    ),
    (
        "Q-X-004",
        "MVP",
        "横向能力",
        "任务详情统一",
        "统一布局/子页与组件",
        "task_detail 系列重构与规范",
        "人天",
        None,
        None,
        "见 docs/03_任务详情",
        "多阶段详情交互一致",
    ),
    # 集成
    (
        "Q-INT-001",
        "MVP",
        "集成与对接",
        "商城订单",
        "鑫采商城订单集成（视接口成熟度）",
        "REST/回调、幂等与对账",
        "人天",
        None,
        None,
        "可单列人天包或按接口点数",
        "订单可进入本系统订单管理",
    ),
    (
        "Q-INT-002",
        "MVP",
        "集成与对接",
        "LCA 计算",
        "手工摆渡方案与接口预留",
        "状态同步、文件/结果传递设计",
        "人天",
        None,
        None,
        "全自动对接可另立二期项",
        "当前阶段可演示人工闭环",
    ),
    (
        "Q-INT-003",
        "二阶段",
        "集成与对接",
        "擎工 SaaS",
        "任务下发、状态回调、结果回传",
        "连接器与接口实现",
        "人天",
        None,
        None,
        "见 11_接口与互操作设计 §5",
        "按二期技术方案验收",
    ),
    (
        "Q-INT-004",
        "MVP",
        "集成与对接",
        "中心连接器",
        "基础版框架（非生产）",
        "服务注册与代理雏形",
        "人天",
        None,
        None,
        "见 11 与第一阶段设计方案 §六",
        "代码交付；标注不用于生产",
    ),
    (
        "Q-INT-005",
        "二阶段",
        "集成与对接",
        "边端与混合流程",
        "轻量化部署与编排（如适用）",
        "见 11 §6",
        "人天",
        None,
        None,
        "第三阶段视评审",
        "按三期方案验收",
    ),
    # 合规
    (
        "Q-STD-001",
        "MVP",
        "标准与合规",
        "TDS/NDI 对齐",
        "架构与预留：身份、目录、合约、审计",
        "方案文档+接口占位+评审支持",
        "人天",
        None,
        None,
        "见 docs/05 与 11、22",
        "设计与实现对齐已确认清单",
    ),
    # PM
    (
        "Q-PM-001",
        "MVP",
        "项目管理与非功能",
        "工程与运维",
        "部署、CI、备份、基础监控",
        "环境与文档",
        "人天",
        None,
        None,
        "",
        "可按环境交付运行说明",
    ),
    (
        "Q-PM-002",
        "MVP",
        "项目管理与非功能",
        "质量保障",
        "测试计划、用例、验收支持",
        "测试报告",
        "人天",
        None,
        None,
        "",
        "核心链路用例通过",
    ),
]

_DAY_ALLOC = _build_day_alloc()
assert _labor_yuan(_DAY_ALLOC) == MVP_LABOR_DAYS * LABOR_RATE + PHASE2_OPT_LABOR_DAYS * LABOR_RATE

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="微软雅黑", size=10.5)
TITLE_MAIN_FONT = Font(bold=True, name="微软雅黑", size=18)


def _phase_label(phase: str) -> str:
    return {
        "MVP": "第一阶段交付",
        "二阶段": "第二阶段交付",
        "可选": "可选范围",
    }.get(phase, phase)


def _sorted_export_rows():
    """SpreadJS 置于末行，其余按应用模块、二级模块、功能点排序。"""
    rows = [r for r in DETAIL_ROWS if r[0] != "Q-X-002"]
    spread = [r for r in DETAIL_ROWS if r[0] == "Q-X-002"]
    assert len(spread) == 1
    rows.sort(key=lambda r: (QID_MXX[r[0]], r[3], r[4], r[0]))
    return rows + spread


def _apply_group_merges(ws, first_row: int, export_rows: list):
    """与子模块、一级功能列一致的纵向合并（对齐招采 xlsx）。"""
    n_data = len(export_rows)
    last_row = first_row + n_data - 1
    wrap_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_t = Alignment(vertical="top", wrap_text=True)
    subs = [MXX_TITLE[QID_MXX[r[0]]] for r in export_rows]
    lev1s = [r[3] for r in export_rows]
    # 子模块（列 1）
    i = 0
    while i < n_data:
        er0 = first_row + i
        sub = subs[i]
        j = i
        while j + 1 < n_data and subs[j + 1] == sub:
            j += 1
        er1 = first_row + j
        if er1 > er0:
            ws.merge_cells(start_row=er0, start_column=1, end_row=er1, end_column=1)
        ws.cell(er0, 1).alignment = wrap_c
        i = j + 1
    # 一级功能（列 2）
    i = 0
    while i < n_data:
        er0 = first_row + i
        sub = subs[i]
        lev1 = lev1s[i]
        j = i
        while j + 1 < n_data and subs[j + 1] == sub and lev1s[j + 1] == lev1:
            j += 1
        er1 = first_row + j
        if er1 > er0 and lev1 is not None:
            ws.merge_cells(start_row=er0, start_column=2, end_row=er1, end_column=2)
        if ws.cell(er0, 2).value is not None:
            ws.cell(er0, 2).alignment = wrap_c
        i = j + 1

    for r in range(first_row, last_row + 1):
        for c in (3, 4, 5, 6):
            ws.cell(r, c).alignment = wrap_t


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "碳足迹数据服务系统"
    fmt_cny = '"¥"#,##0;_-* "-"??_-;_-@_-'

    # 与招采参考单一致的 6 列 + 顶部标题行
    title_row, header_row, data_row0 = 1, 2, 3

    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=6)
    tcell = ws.cell(title_row, 1)
    tcell.value = (
        "钢铁行业产业链碳足迹数据服务系统\n"
        "拆分工作量评估与报价清单（人天单价 1500 元；SpreadJS 为固定价）"
    )
    tcell.font = TITLE_MAIN_FONT
    tcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[title_row].height = 44

    headers = [
        "子模块",
        "一级功能",
        "二级功能",
        "预估工作量(人天)",
        "单价",
        "预估报价(元)",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    export_rows = _sorted_export_rows()
    n = len(export_rows)
    for i, row in enumerate(export_rows):
        r = data_row0 + i
        qid, phase, _m1, m2, feat, _deliver, *_rest = row
        sub = MXX_TITLE[QID_MXX[qid]]
        lev2 = feat if phase == "MVP" else f"{feat}（{_phase_label(phase)}）"
        prev = export_rows[i - 1] if i else None
        prev_sub = MXX_TITLE[QID_MXX[prev[0]]] if prev else None
        prev_m2 = prev[3] if prev else None
        ws.cell(r, 1, value=sub if prev_sub != sub else None)
        ws.cell(r, 2, value=m2 if (prev_sub != sub or prev_m2 != m2) else None)
        ws.cell(r, 3, value=lev2)
        if qid == "Q-X-002":
            ws.cell(r, 4, value=None)
            ws.cell(r, 5, value=None)
            ws.cell(r, 6, value=SPREADJS_FIXED_YUAN)
        else:
            d = _DAY_ALLOC[qid]
            ws.cell(r, 4, value=d)
            ws.cell(r, 5, value=LABOR_RATE)
            ws.cell(r, 6, value=f"=D{r}*E{r}")
        ws.cell(r, 4).number_format = "#,##0"
        ws.cell(r, 5).number_format = "#,##0"
        ws.cell(r, 6).number_format = fmt_cny

    _apply_group_merges(ws, data_row0, export_rows)

    last_data = data_row0 + n - 1
    tr = last_data + 1
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=3)
    ws.cell(tr, 1, value="总计")
    ws.cell(tr, 1).font = Font(bold=True, name="微软雅黑")
    ws.cell(tr, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(tr, 4, value=f"=SUM(D{data_row0}:D{last_data})")
    ws.cell(tr, 4).font = Font(bold=True, name="微软雅黑")
    ws.cell(tr, 4).number_format = "#,##0"
    ws.cell(tr, 5, value=LABOR_RATE)
    ws.cell(tr, 5).number_format = "#,##0"
    ws.cell(tr, 6, value=f"=SUM(F{data_row0}:F{last_data})")
    ws.cell(tr, 6).font = Font(bold=True, name="微软雅黑")
    ws.cell(tr, 6).number_format = fmt_cny

    nr = tr + 1
    ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=6)
    total_days = MVP_LABOR_DAYS + PHASE2_OPT_LABOR_DAYS
    ws.cell(nr, 1).value = (
        f"编制说明：与《招采电子交易平台分布式「招采通」拆分工作量评估与报价清单》列结构一致；"
        f"工作量单位为「人天」，整数；人天行单价 {LABOR_RATE} 元/人天，人工合计 {total_days} 人天、"
        f"小计 {total_days * LABOR_RATE} 元；GrapeCity SpreadJS 商业授权按固定价 {SPREADJS_FIXED_YUAN} 元单列（不计入上方人天合计）；"
        f"报价总计 {TOTAL_QUOTE_YUAN} 元。"
    )
    ws.cell(nr, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[nr].height = 60

    widths = [28, 18, 44, 12, 10, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{data_row0}"
    wb.save(OUT)
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
